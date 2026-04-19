import json
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter


WORKBOOK_PATH = Path("orders2026.xlsx")
OUTPUT_PATH = Path("shipments-workbook-data.js")
MONTH_ORDER = ["январь 2026", "февраль 2026", "март 2026", "апрель 2026"]


def normalize_title(title: str) -> str:
    return str(title or "").strip().lower()


def format_value(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, date):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def fill_color(cell) -> str:
    fill = cell.fill
    if not fill or fill.fill_type != "solid":
        return ""
    rgb = fill.fgColor.rgb or fill.start_color.rgb or ""
    rgb = str(rgb or "")
    if not rgb or rgb in {"00000000", "000000", "0000000", "FFFFFFFF"}:
        return ""
    if len(rgb) == 8:
        rgb = rgb[2:]
    if len(rgb) != 6:
        return ""
    return f"#{rgb.lower()}"


def alignment(cell) -> str:
    value = cell.alignment.horizontal or ""
    return "" if value == "general" else value


def get_used_range(ws):
    max_row = 0
    max_col = 20
    for row_idx in range(1, ws.max_row + 1):
        has_content = False
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row_idx, col_idx)
            if isinstance(cell, MergedCell):
                continue
            if cell.value not in (None, "") or fill_color(cell):
                has_content = True
                break
        if has_content:
            max_row = row_idx
    return max_row, max_col


def merged_maps(ws):
    top_left = {}
    covered = set()
    for merged in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged.bounds
        top_left[(min_row, min_col)] = {
            "rowspan": max_row - min_row + 1,
            "colspan": max_col - min_col + 1,
        }
        for row_idx in range(min_row, max_row + 1):
            for col_idx in range(min_col, max_col + 1):
                if row_idx == min_row and col_idx == min_col:
                    continue
                covered.add((row_idx, col_idx))
    return top_left, covered


def get_column_widths(ws, max_col):
    widths = []
    for col_idx in range(1, max_col + 1):
        dim = ws.column_dimensions.get(get_column_letter(col_idx))
        width = dim.width if dim and dim.width else 12
        widths.append(round(width * 7.2, 1))
    return widths


def get_row_height(ws, row_idx):
    dim = ws.row_dimensions.get(row_idx)
    return dim.height if dim and dim.height else None


def serialize_sheet(ws_values, ws_styles):
    max_row, max_col = get_used_range(ws_values)
    merged_top_left, merged_covered = merged_maps(ws_styles)
    rows = []

    for row_idx in range(1, max_row + 1):
        cells = []
        for col_idx in range(1, max_col + 1):
            if (row_idx, col_idx) in merged_covered:
                continue
            style_cell = ws_styles.cell(row_idx, col_idx)
            value_cell = ws_values.cell(row_idx, col_idx)
            merged_meta = merged_top_left.get((row_idx, col_idx), {"rowspan": 1, "colspan": 1})
            cells.append(
                {
                    "value": format_value(value_cell.value),
                    "bg": fill_color(style_cell),
                    "bold": bool(style_cell.font.bold),
                    "align": alignment(style_cell),
                    "rowspan": merged_meta["rowspan"],
                    "colspan": merged_meta["colspan"],
                }
            )
        rows.append({"height": get_row_height(ws_styles, row_idx), "cells": cells})

    return {
        "title": ws_values.title,
        "rows": rows,
        "colWidths": get_column_widths(ws_styles, max_col),
        "maxCol": max_col,
        "maxRow": max_row,
    }


def main():
    wb_values = load_workbook(WORKBOOK_PATH, data_only=True)
    wb_styles = load_workbook(WORKBOOK_PATH, data_only=False)
    title_map = {normalize_title(ws.title): ws.title for ws in wb_values.worksheets}
    ordered_titles = [title_map[name] for name in MONTH_ORDER if name in title_map]

    payload = {"sheets": []}
    for title in ordered_titles:
        payload["sheets"].append(serialize_sheet(wb_values[title], wb_styles[title]))

    OUTPUT_PATH.write_text(
        "window.SHIPMENTS_WORKBOOK = " + json.dumps(payload, ensure_ascii=False) + ";\n",
        encoding="utf-8",
    )
    print(f"Exported {len(payload['sheets'])} sheets to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()